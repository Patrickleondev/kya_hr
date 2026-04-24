# -*- coding: utf-8 -*-
"""Import de présences depuis un fichier Excel.

Colonnes supportées (insensible à la casse, accents/espaces ignorés) :
- matricule (prioritaire pour retrouver l'employé)
- nom / employe / employee_name (fallback)
- date (obligatoire, formats : YYYY-MM-DD, DD/MM/YYYY, Excel date)
- statut / status : Present / Absent / Half Day / On Leave / Work From Home
- check_in / arrivee / heure_entree (optionnel)
- check_out / depart / heure_sortie (optionnel)
- retard (optionnel, booléen)
- departement / department (optionnel)
"""
from __future__ import annotations

import frappe
from frappe import _
from frappe.utils import getdate, get_site_path
from frappe.utils.file_manager import get_file_path

import os
import unicodedata


STATUS_MAP = {
    "present": "Present", "p": "Present", "présent": "Present", "presente": "Present",
    "absent": "Absent", "a": "Absent", "abs": "Absent",
    "half day": "Half Day", "demi": "Half Day", "demi-journée": "Half Day", "hd": "Half Day",
    "on leave": "On Leave", "congé": "On Leave", "conge": "On Leave", "l": "On Leave",
    "work from home": "Work From Home", "télétravail": "Work From Home", "wfh": "Work From Home",
}


def _norm(s):
    if s is None:
        return ""
    s = str(s).strip().lower()
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
    return s.replace("_", " ").replace("-", " ").strip()


def _parse_date(val):
    if not val and val != 0:
        return None
    if hasattr(val, "strftime"):
        return val
    # Excel numeric date
    if isinstance(val, (int, float)):
        from datetime import datetime, timedelta
        try:
            return (datetime(1899, 12, 30) + timedelta(days=float(val))).date()
        except Exception:
            return None
    s = str(val).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            from datetime import datetime
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    try:
        return getdate(s)
    except Exception:
        return None


def _find_employee(matricule, name):
    """Trouve un employé par matricule (custom field) puis par nom."""
    if matricule:
        emp = (
            frappe.db.get_value("Employee", {"custom_matricule_kya": matricule}, "name")
            or frappe.db.get_value("Employee", {"employee_number": matricule}, "name")
        )
        if emp:
            return emp
    if name:
        emp = frappe.db.get_value("Employee", {"employee_name": name}, "name")
        if emp:
            return emp
        # recherche floue
        rows = frappe.db.sql(
            "SELECT name FROM `tabEmployee` WHERE employee_name LIKE %s LIMIT 1",
            ("%" + str(name).strip() + "%",),
        )
        if rows:
            return rows[0][0]
    return None


@frappe.whitelist()
def import_attendance(file_url: str):
    """Importe les présences depuis un Excel uploadé.

    Retourne : {inserted, updated, skipped, errors[]}.
    """
    if not frappe.has_permission("Attendance", "create"):
        frappe.throw(_("Permission refusée : création d'Attendance."), frappe.PermissionError)

    # Résoudre chemin fichier
    path = get_file_path(file_url) if file_url.startswith("/files/") or file_url.startswith("/private/") else file_url
    if not os.path.isfile(path):
        frappe.throw(_("Fichier introuvable : {0}").format(file_url))

    # Lire Excel via openpyxl
    try:
        from openpyxl import load_workbook
    except ImportError:
        frappe.throw(_("openpyxl non installé sur le site."))

    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows or len(rows) < 2:
        frappe.throw(_("Fichier vide ou sans ligne de données."))

    # Header → map colonnes
    headers = [_norm(c) for c in rows[0]]
    def col(*aliases):
        for a in aliases:
            a = _norm(a)
            if a in headers:
                return headers.index(a)
        return -1

    idx_mat = col("matricule", "matricule employe", "matricul")
    idx_nom = col("nom", "employe", "employee name", "employee", "nom employe", "prenom nom")
    idx_date = col("date", "jour", "attendance date")
    idx_status = col("statut", "status", "etat", "presence")
    idx_in = col("check in", "arrivee", "heure entree", "in time")
    idx_out = col("check out", "depart", "heure sortie", "out time")
    idx_dept = col("departement", "department", "service")

    if idx_date == -1:
        frappe.throw(_("Colonne 'date' introuvable dans l'en-tête. Colonnes détectées : {0}").format(", ".join(headers)))

    inserted, updated, skipped = 0, 0, 0
    errors = []

    for i, r in enumerate(rows[1:], start=2):
        try:
            matricule = str(r[idx_mat]).strip() if idx_mat != -1 and r[idx_mat] is not None else ""
            nom = str(r[idx_nom]).strip() if idx_nom != -1 and r[idx_nom] is not None else ""
            if not matricule and not nom:
                skipped += 1
                continue
            emp = _find_employee(matricule, nom)
            if not emp:
                errors.append(f"L{i} : employé introuvable (matricule={matricule!r}, nom={nom!r})")
                skipped += 1
                continue

            att_date = _parse_date(r[idx_date])
            if not att_date:
                errors.append(f"L{i} : date invalide ({r[idx_date]!r})")
                skipped += 1
                continue

            status_raw = _norm(r[idx_status]) if idx_status != -1 and r[idx_status] is not None else "present"
            status = STATUS_MAP.get(status_raw, "Present")

            check_in = r[idx_in] if idx_in != -1 else None
            check_out = r[idx_out] if idx_out != -1 else None

            # Existe déjà ?
            existing = frappe.db.get_value(
                "Attendance", {"employee": emp, "attendance_date": att_date, "docstatus": ["<", 2]}, "name"
            )
            if existing:
                doc = frappe.get_doc("Attendance", existing)
                doc.status = status
                if check_in:
                    doc.in_time = check_in
                if check_out:
                    doc.out_time = check_out
                doc.save(ignore_permissions=False)
                updated += 1
            else:
                doc = frappe.new_doc("Attendance")
                doc.employee = emp
                doc.attendance_date = att_date
                doc.status = status
                if check_in:
                    doc.in_time = check_in
                if check_out:
                    doc.out_time = check_out
                doc.company = frappe.db.get_value("Employee", emp, "company") or frappe.defaults.get_user_default("Company")
                doc.insert(ignore_permissions=False)
                try:
                    doc.submit()
                except Exception:
                    pass
                inserted += 1
        except Exception as e:
            errors.append(f"L{i} : {e}")
            skipped += 1

    frappe.db.commit()
    return {
        "inserted": inserted,
        "updated": updated,
        "skipped": skipped,
        "errors": errors[:50],
        "total_errors": len(errors),
    }
