# -*- coding: utf-8 -*-
"""Import/mise à jour des matricules KYA sur les employés depuis un Excel."""
from __future__ import annotations

import os
import unicodedata

import frappe
from frappe import _
from frappe.utils.file_manager import get_file_path


def _norm(s):
    if s is None:
        return ""
    s = str(s).strip().lower()
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
    return s.replace("_", " ").replace("-", " ").strip()


def _find_employee(nom: str, email: str = ""):
    """Trouve un employé par email (prioritaire) puis par nom exact/approchant."""
    if email:
        emp = frappe.db.get_value("Employee", {"personal_email": email}, "name") or \
              frappe.db.get_value("Employee", {"company_email": email}, "name") or \
              frappe.db.get_value("Employee", {"user_id": email}, "name")
        if emp:
            return emp
    if nom:
        emp = frappe.db.get_value("Employee", {"employee_name": nom}, "name")
        if emp:
            return emp
        rows = frappe.db.sql(
            "SELECT name FROM `tabEmployee` WHERE employee_name LIKE %s LIMIT 1",
            ("%" + nom + "%",),
        )
        if rows:
            return rows[0][0]
    return None


@frappe.whitelist()
def import_matricules(file_url: str):
    """Importe/met à jour `custom_matricule_kya` depuis un Excel.

    Colonnes détectées : matricule, nom (ou employe/employee_name), email (optionnel).
    """
    if not frappe.has_permission("Employee", "write"):
        frappe.throw(_("Permission refusée."), frappe.PermissionError)

    path = get_file_path(file_url) if file_url.startswith(("/files/", "/private/")) else file_url
    if not os.path.isfile(path):
        frappe.throw(_("Fichier introuvable : {0}").format(file_url))

    try:
        from openpyxl import load_workbook
    except ImportError:
        frappe.throw(_("openpyxl non installé."))

    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        frappe.throw(_("Fichier vide."))

    headers = [_norm(c) for c in rows[0]]

    def col(*aliases):
        for a in aliases:
            na = _norm(a)
            if na in headers:
                return headers.index(na)
        return -1

    idx_mat = col("matricule", "matricule kya", "custom matricule kya", "id")
    idx_nom = col("nom", "employe", "employee name", "nom complet", "prenom nom", "nom et prenom")
    idx_email = col("email", "mail", "e mail", "company email")

    if idx_mat == -1:
        frappe.throw(_("Colonne 'matricule' introuvable. Colonnes : {0}").format(", ".join(headers)))
    if idx_nom == -1 and idx_email == -1:
        frappe.throw(_("Aucune colonne d'identification (nom ou email) trouvée."))

    updated, created, skipped = 0, 0, 0
    errors = []

    for i, r in enumerate(rows[1:], start=2):
        try:
            mat = str(r[idx_mat]).strip() if r[idx_mat] is not None else ""
            nom = str(r[idx_nom]).strip() if idx_nom != -1 and r[idx_nom] is not None else ""
            email = str(r[idx_email]).strip() if idx_email != -1 and r[idx_email] is not None else ""
            if not mat:
                skipped += 1
                continue
            emp = _find_employee(nom, email)
            if not emp:
                errors.append(f"L{i} : employé introuvable (nom={nom!r}, email={email!r})")
                skipped += 1
                continue
            current = frappe.db.get_value("Employee", emp, "custom_matricule_kya")
            if current == mat:
                skipped += 1
                continue
            frappe.db.set_value("Employee", emp, "custom_matricule_kya", mat, update_modified=False)
            updated += 1
        except Exception as e:
            errors.append(f"L{i} : {e}")
            skipped += 1

    frappe.db.commit()
    return {"updated": updated, "created": created, "skipped": skipped,
            "errors": errors[:50], "total_errors": len(errors)}


@frappe.whitelist(allow_guest=True)
def lookup_by_matricule(matricule: str):
    """Recherche d'un employé par matricule — utilisé par les KYA Forms publics.

    Retour sans info sensible (juste nom, dépt, poste, matricule).
    """
    if not matricule:
        return None
    mat = str(matricule).strip()
    emp = frappe.db.get_value(
        "Employee",
        {"custom_matricule_kya": mat, "status": "Active"},
        ["name", "employee_name", "department", "designation", "custom_matricule_kya", "employment_type"],
        as_dict=True,
    )
    if not emp:
        return None
    return {
        "employee": emp.name,
        "employee_name": emp.employee_name,
        "department": emp.department,
        "designation": emp.designation,
        "matricule": emp.custom_matricule_kya,
        "employment_type": emp.employment_type,
    }
