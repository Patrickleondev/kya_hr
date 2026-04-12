import frappe
from frappe.utils import today, now_datetime, add_days, getdate, get_datetime, cint


# ΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉ
#  QUICK LINKS (HRMS Mobile)
# ΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉ

# Types d'emploi consid├⌐r├⌐s comme "personnel interne"
EMPLOYEE_TYPES = {"CDI", "CDD", "Full-time", "Part-time", "Probation", "Contract"}
# Types consid├⌐r├⌐s comme "stagiaires"
STAGE_TYPES = {"Stage", "Intern", "Apprentice"}
# Types consid├⌐r├⌐s comme "prestataires"
PRESTATAIRE_TYPES = {"Prestataire de service"}


def _get_user_category(employment_type):
    """D├⌐termine la cat├⌐gorie utilisateur ├á partir du type d'emploi."""
    if employment_type in STAGE_TYPES:
        return "stage"
    elif employment_type in PRESTATAIRE_TYPES:
        return "prestataire"
    else:
        return "employee"


@frappe.whitelist()
def get_kya_quick_links():
    """Return KYA-specific quick links based on current user's employment type."""
    user = frappe.session.user
    if user == "Guest":
        return []

    employee = frappe.db.get_value(
        "Employee",
        {"user_id": user, "status": "Active"},
        ["name", "employment_type", "employee_name"],
        as_dict=True,
    )

    if not employee:
        return []

    category = _get_user_category(employee.get("employment_type") or "")
    links = []

    if category == "stage":
        # --- Stagiaire links ---
        links.append({
            "title": "Permission de Sortie",
            "description": "Demander une permission de sortie",
            "url": "/permission-sortie-stagiaire/new",
            "emoji": "≡ƒÜ¬",
        })
        links.append({
            "title": "Bilan de Stage",
            "description": "Remplir le bilan de fin de stage",
            "url": "/bilan-fin-de-stage/new",
            "emoji": "≡ƒô¥",
        })
    elif category == "prestataire":
        # --- Prestataire de service links ---
        links.append({
            "title": "Permission de Sortie",
            "description": "Demander une permission de sortie",
            "url": "/permission-sortie-employe/new",
            "emoji": "≡ƒÜ¬",
        })
    else:
        # --- Employee (CDI/CDD) links ---
        links.append({
            "title": "Permission de Sortie",
            "description": "Demander une permission de sortie",
            "url": "/permission-sortie-employe/new",
            "emoji": "≡ƒÜ¬",
        })
        links.append({
            "title": "Planning de Cong├⌐",
            "description": "Planifier vos p├⌐riodes de cong├⌐ annuelles",
            "url": "/planning-conge/new",
            "emoji": "≡ƒôà",
        })

    # --- Common links (all types) ---
    links.append({
        "title": "PV Sortie Mat├⌐riel",
        "description": "D├⌐clarer une sortie de mat├⌐riel",
        "url": "/pv-sortie-materiel/new",
        "emoji": "≡ƒôª",
    })
    links.append({
        "title": "Demande d\'Achat",
        "description": "Soumettre une demande d\'achat",
        "url": "/demande-achat/new",
        "emoji": "≡ƒ¢Æ",
    })

    return links


@frappe.whitelist()
def get_user_category():
    """Return the current user's category (stage/prestataire/employee).
    Used by the HRMS frontend to apply UI filters."""
    user = frappe.session.user
    if user == "Guest":
        return {"category": "guest", "employment_type": None}

    emp = frappe.db.get_value(
        "Employee",
        {"user_id": user, "status": "Active"},
        ["name", "employment_type"],
        as_dict=True,
    )
    if not emp:
        return {"category": "unknown", "employment_type": None}

    return {
        "category": _get_user_category(emp.employment_type or ""),
        "employment_type": emp.employment_type,
    }


# ΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉ
#  WORKFLOW ACTIONS (Web Form Mobile Approval)
# ΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉ

ALLOWED_DOCTYPES = {
    "Permission Sortie Stagiaire",
    "Permission Sortie Employe",
    "PV Sortie Materiel",
    "Planning Conge",
    "Demande Achat KYA",
    "Bilan Fin de Stage",
}


@frappe.whitelist()
def get_kya_workflow_actions(doctype, docname):
    """Get available workflow actions for the current user on a document.
    Returns current workflow_state and list of possible actions."""
    if doctype not in ALLOWED_DOCTYPES:
        frappe.throw("Type de document non autorisé", frappe.PermissionError)

    doc = frappe.get_doc(doctype, docname)
    doc.check_permission("read")

    from frappe.model.workflow import get_transitions

    transitions = get_transitions(doc, raise_exception=False)
    actions = []
    for t in transitions:
        actions.append({
            "action": t.get("action"),
            "next_state": t.get("next_state"),
        })

    return {
        "workflow_state": doc.workflow_state,
        "docstatus": doc.docstatus,
        "actions": actions,
    }


@frappe.whitelist()
def apply_kya_workflow_action(doctype, docname, action):
    """Apply a workflow action from the web form (mobile approval)."""
    if doctype not in ALLOWED_DOCTYPES:
        frappe.throw("Type de document non autorisé", frappe.PermissionError)

    from frappe.model.workflow import apply_workflow

    doc = frappe.get_doc(doctype, docname)
    doc.check_permission("write")
    apply_workflow(doc, action)

    return {
        "status": "success",
        "workflow_state": doc.workflow_state,
        "docstatus": doc.docstatus,
    }


# ΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉ
#  BIOMETRIC ATTENDANCE API (for frontend dashboard)
# ΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉ

@frappe.whitelist()
def get_attendance_summary(from_date=None, to_date=None, department=None):
    """Get attendance summary (present, absent, late, half day) for a date range."""
    if not from_date:
        from_date = getdate(today()).replace(day=1).isoformat()
    if not to_date:
        to_date = today()

    filters = {
        "attendance_date": ["between", [from_date, to_date]],
        "docstatus": 1,
    }
    if department:
        filters["department"] = department

    summary = frappe.db.sql("""
        SELECT status, COUNT(*) as count
        FROM `tabAttendance`
        WHERE attendance_date BETWEEN %(from_date)s AND %(to_date)s
          AND docstatus = 1
          {dept_filter}
        GROUP BY status
    """.format(
        dept_filter="AND department = %(department)s" if department else ""
    ), {
        "from_date": from_date,
        "to_date": to_date,
        "department": department,
    }, as_dict=True)

    daily = frappe.db.sql("""
        SELECT attendance_date, status, COUNT(*) as count
        FROM `tabAttendance`
        WHERE attendance_date BETWEEN %(from_date)s AND %(to_date)s
          AND docstatus = 1
          {dept_filter}
        GROUP BY attendance_date, status
        ORDER BY attendance_date
    """.format(
        dept_filter="AND department = %(department)s" if department else ""
    ), {
        "from_date": from_date,
        "to_date": to_date,
        "department": department,
    }, as_dict=True)

    emp_filters = {"status": "Active"}
    if department:
        emp_filters["department"] = department
    total_employees = frappe.db.count("Employee", emp_filters)

    departments = frappe.db.sql_list(
        "SELECT DISTINCT department FROM `tabEmployee` WHERE status='Active' AND department IS NOT NULL"
    )

    return {
        "from_date": from_date,
        "to_date": to_date,
        "total_employees": total_employees,
        "summary": {row.status: row.count for row in summary},
        "daily_breakdown": daily,
        "departments": departments,
    }


@frappe.whitelist()
def get_employee_checkins(employee=None, from_date=None, to_date=None,
                          department=None, limit_page_length=100, limit_start=0):
    """Get Employee Checkin logs for attendance tracking."""
    if not from_date:
        from_date = today()
    if not to_date:
        to_date = today()

    filters = {
        "time": ["between", [
            f"{from_date} 00:00:00",
            f"{to_date} 23:59:59",
        ]],
    }
    if employee:
        filters["employee"] = employee
    if department:
        filters["department"] = department

    checkins = frappe.get_all(
        "Employee Checkin",
        filters=filters,
        fields=[
            "name", "employee", "employee_name", "department",
            "time", "log_type", "device_id",
            "attendance", "shift",
        ],
        order_by="time desc",
        limit_page_length=int(limit_page_length),
        limit_start=int(limit_start),
    )

    total = frappe.db.count("Employee Checkin", filters)

    return {
        "data": checkins,
        "total": total,
        "from_date": from_date,
        "to_date": to_date,
    }


@frappe.whitelist()
def get_attendance_dashboard(date=None, department=None):
    """Get real-time attendance dashboard data for a specific date."""
    if not date:
        date = today()

    emp_filters = {"status": "Active"}
    if department:
        emp_filters["department"] = department

    all_employees = frappe.get_all(
        "Employee",
        filters=emp_filters,
        fields=["name", "employee_name", "department", "designation", "image"],
    )
    emp_ids = [e.name for e in all_employees]
    emp_map = {e.name: e for e in all_employees}

    att_records = frappe.get_all(
        "Attendance",
        filters={
            "attendance_date": date,
            "docstatus": 1,
            "employee": ["in", emp_ids] if emp_ids else ["in", ["__none__"]],
        },
        fields=["employee", "status", "leave_type", "late_entry", "early_exit"],
    )
    att_map = {a.employee: a for a in att_records}

    checkins = frappe.db.sql("""
        SELECT employee,
               MIN(CASE WHEN log_type = 'IN' THEN time END) as first_in,
               MAX(CASE WHEN log_type = 'OUT' THEN time END) as last_out,
               COUNT(*) as total_punches
        FROM `tabEmployee Checkin`
        WHERE DATE(time) = %(date)s
          AND employee IN %(employees)s
        GROUP BY employee
    """, {
        "date": date,
        "employees": emp_ids or ["__none__"],
    }, as_dict=True) if emp_ids else []
    checkin_map = {c.employee: c for c in checkins}

    present = []
    absent = []
    on_leave = []
    late = []

    for emp_id in emp_ids:
        emp = emp_map[emp_id]
        att = att_map.get(emp_id)
        ci = checkin_map.get(emp_id)

        entry = {
            "employee": emp.name,
            "employee_name": emp.employee_name,
            "department": emp.department,
            "designation": emp.designation,
            "image": emp.image,
            "first_in": str(ci.first_in) if ci and ci.first_in else None,
            "last_out": str(ci.last_out) if ci and ci.last_out else None,
            "total_punches": ci.total_punches if ci else 0,
        }

        if att:
            entry["status"] = att.status
            entry["leave_type"] = att.leave_type
            entry["late_entry"] = att.late_entry
            entry["early_exit"] = att.early_exit

            if att.status == "On Leave":
                on_leave.append(entry)
            elif att.status in ("Present", "Half Day", "Work From Home"):
                present.append(entry)
                if att.late_entry:
                    late.append(entry)
            else:
                absent.append(entry)
        else:
            entry["status"] = "Absent"
            absent.append(entry)

    return {
        "date": date,
        "department": department,
        "stats": {
            "total": len(emp_ids),
            "present": len(present),
            "absent": len(absent),
            "on_leave": len(on_leave),
            "late": len(late),
        },
        "present": present,
        "absent": absent,
        "on_leave": on_leave,
        "late": late,
    }


@frappe.whitelist()
def get_department_attendance_stats(from_date=None, to_date=None):
    """Get attendance breakdown by department for a date range."""
    if not from_date:
        from_date = getdate(today()).replace(day=1).isoformat()
    if not to_date:
        to_date = today()

    stats = frappe.db.sql("""
        SELECT
            e.department,
            a.status,
            COUNT(*) as count
        FROM `tabAttendance` a
        JOIN `tabEmployee` e ON e.name = a.employee
        WHERE a.attendance_date BETWEEN %(from_date)s AND %(to_date)s
          AND a.docstatus = 1
          AND e.department IS NOT NULL
        GROUP BY e.department, a.status
        ORDER BY e.department
    """, {"from_date": from_date, "to_date": to_date}, as_dict=True)

    dept_map = {}
    for row in stats:
        dept = row.department
        if dept not in dept_map:
            dept_map[dept] = {"department": dept}
        dept_map[dept][row.status] = row.count

    return {
        "from_date": from_date,
        "to_date": to_date,
        "data": list(dept_map.values()),
    }


@frappe.whitelist()
def get_employee_attendance_detail(employee, from_date=None, to_date=None):
    """Get detailed attendance for a specific employee."""
    if not from_date:
        from_date = getdate(today()).replace(day=1).isoformat()
    if not to_date:
        to_date = today()

    emp = frappe.get_value(
        "Employee", employee,
        ["name", "employee_name", "department", "designation",
         "attendance_device_id", "image", "company"],
        as_dict=True,
    )
    if not emp:
        frappe.throw("Employ├⌐ non trouv├⌐", frappe.DoesNotExistError)

    attendance = frappe.get_all(
        "Attendance",
        filters={
            "employee": employee,
            "attendance_date": ["between", [from_date, to_date]],
            "docstatus": 1,
        },
        fields=["attendance_date", "status", "leave_type",
                "late_entry", "early_exit", "working_hours"],
        order_by="attendance_date desc",
    )

    checkins = frappe.get_all(
        "Employee Checkin",
        filters={
            "employee": employee,
            "time": ["between", [
                f"{from_date} 00:00:00",
                f"{to_date} 23:59:59",
            ]],
        },
        fields=["time", "log_type", "device_id", "shift"],
        order_by="time desc",
    )

    summary = {}
    for a in attendance:
        status = a.status
        summary[status] = summary.get(status, 0) + 1

    return {
        "employee": emp,
        "from_date": from_date,
        "to_date": to_date,
        "attendance": attendance,
        "checkins": checkins,
        "summary": summary,
    }


# ΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉ
#  MES DEMANDES (Document Tracking for Mobile)
# ΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉ

_TRACKABLE_DOCTYPES = [
    {
        "doctype": "Permission Sortie Employe",
        "label": "Permission de Sortie",
        "category": "rh",
        "for": ["employee", "prestataire"],
        "route": "permission-sortie-employe",
    },
    {
        "doctype": "Permission Sortie Stagiaire",
        "label": "Permission de Sortie",
        "category": "rh",
        "for": ["stage"],
        "route": "permission-sortie-stagiaire",
    },
    {
        "doctype": "Planning Conge",
        "label": "Planning de Cong\u00e9",
        "category": "rh",
        "for": ["employee"],
        "route": "planning-conge",
    },
    {
        "doctype": "PV Sortie Materiel",
        "label": "PV Sortie Mat\u00e9riel",
        "category": "stock",
        "for": ["employee", "stage", "prestataire"],
        "filter_by": "owner",  # No Employee Link field, filter by doc creator
        "route": "pv-sortie-materiel",
    },
    {
        "doctype": "Demande Achat KYA",
        "label": "Demande d'Achat",
        "category": "achats",
        "for": ["employee"],
        "route": "demande-achat",
    },
    {
        "doctype": "Bilan Fin de Stage",
        "label": "Bilan de Stage",
        "category": "rh",
        "for": ["stage"],
        "route": "bilan-fin-de-stage",
    },
]

_STATE_COLORS = {
    "Brouillon": "orange",
    "En attente Chef": "blue",
    "En attente Chef Service": "blue",
    "En attente RH": "blue",
    "En attente DG": "blue",
    "En attente DGA": "blue",
    "En attente Magasin": "blue",
    "En attente Audit": "blue",
        "En attente Comptabilité": "blue",
    "En attente Resp. Stagiaires": "blue",
    "En cours": "yellow",
        "Approuvé": "green",
        "Approuvée": "green",
        "Validé": "green",
        "Rejeté": "red",
        "Rejetée": "red",
        "Annulé": "gray",
}


@frappe.whitelist()
def get_my_documents(limit=20, offset=0, status_filter=None):
    """Retourne les documents de l'utilisateur courant avec ├⌐tats workflow."""
    user = frappe.session.user
    if user == "Guest":
        return {"data": [], "total": 0}

    employee = frappe.db.get_value(
        "Employee",
        {"user_id": user, "status": "Active"},
        ["name", "employment_type"],
        as_dict=True,
    )
    if not employee:
        return {"data": [], "total": 0}

    category = _get_user_category(employee.employment_type or "")
    documents = []

    for dt_info in _TRACKABLE_DOCTYPES:
        if category not in dt_info["for"]:
            continue

        dt = dt_info["doctype"]
        if not frappe.db.exists("DocType", dt):
            continue

        filter_field = dt_info.get("filter_by", "employee")
        if filter_field == "owner":
            filters = {"owner": user}
        else:
            filters = {filter_field: employee.name}
        if status_filter and status_filter != "all":
            filters["workflow_state"] = status_filter

        try:
            docs = frappe.get_all(
                dt,
                filters=filters,
                fields=["name", "creation", "modified", "workflow_state", "docstatus"],
                order_by="modified desc",
                limit_page_length=50,
            )
        except Exception:
            continue

        for doc in docs:
            ws = doc.get("workflow_state") or "Brouillon"
            route = dt_info.get("route") or frappe.scrub(dt).replace("_", "-")
            documents.append({
                "doctype": dt,
                "name": doc.name,
                "label": dt_info["label"],
                "category": dt_info["category"],
                "workflow_state": ws,
                "color": _STATE_COLORS.get(ws, "gray"),
                "creation": str(doc.creation),
                "modified": str(doc.modified),
                "url": f"/{route}/{doc.name}",
            })

    # Sort by modified desc
    documents.sort(key=lambda d: d["modified"], reverse=True)
    total = len(documents)
    page = documents[cint(offset): cint(offset) + cint(limit)]

    return {"data": page, "total": total}


# ΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉ
#  ENQU├èTES & ├ëVALUATIONS (KYA Forms for Mobile)
# ΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉ

@frappe.whitelist()
def get_my_kya_forms():
    """Retourne les enqu├¬tes/├⌐valuations disponibles et remplies pour l'utilisateur."""
    user = frappe.session.user
    if user == "Guest":
        return {"available": [], "completed": []}

    # V├⌐rifier si le DocType KYA Form existe (kya_services install├⌐)
    if not frappe.db.exists("DocType", "KYA Form"):
        return {"available": [], "completed": []}

    employee = frappe.db.get_value(
        "Employee",
        {"user_id": user, "status": "Active"},
        ["name", "employee_name", "department"],
        as_dict=True,
    )

    # Formulaires actifs (statut = Actif)
    available = []
    try:
        forms = frappe.get_all(
            "KYA Form",
            filters={"statut": "Actif"},
            fields=["name", "titre", "description", "type_formulaire",
                     "date_limite", "equipe_cible"],
            order_by="creation desc",
        )
        for f in forms:
            # Vérifier si l'employé a déjà une réponse soumise
            already_done = False
            if employee and frappe.db.exists("DocType", "KYA Form Response"):
                already_done = bool(frappe.db.get_value(
                    "KYA Form Response",
                    {"formulaire": f.name, "employe": employee.name, "soumis_le": ["is", "set"]},
                    "name",
                ))
            # Token de réponse personnel pour cet employé
            personal_token = None
            if employee and frappe.db.exists("DocType", "KYA Form Response"):
                personal_token = frappe.db.get_value(
                    "KYA Form Response",
                    {"formulaire": f.name, "employe": employee.name, "soumis_le": ["", None]},
                    "token",
                )
            available.append({
                "name": f.name,
                "titre": f.titre,
                "description": f.description or "",
                "type": f.type_formulaire or "Enquête",
                "date_limite": str(f.date_limite) if f.date_limite else None,
                "equipe_cible": f.equipe_cible or "",
                "completed": already_done,
                "url": f"/kya-survey?token={personal_token}" if personal_token else f"/kya-survey?form={f.name}",
            })
    except Exception:
        pass

    # Réponses déjà soumises par cet employé
    completed = []
    try:
        if employee and frappe.db.exists("DocType", "KYA Form Response"):
            responses = frappe.get_all(
                "KYA Form Response",
                filters={"employe": employee.name, "soumis_le": ["is", "set"]},
                fields=["name", "formulaire", "soumis_le"],
                order_by="soumis_le desc",
                limit_page_length=20,
            )
            for r in responses:
                form_title = frappe.db.get_value("KYA Form", r.formulaire, "titre") or r.formulaire
                completed.append({
                    "name": r.name,
                    "form": r.formulaire,
                    "form_title": form_title,
                    "date": str(r.soumis_le),
                })
    except Exception:
        pass

    return {"available": available, "completed": completed}


# ΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉ
# (Old get_my_tasks removed — replaced by the version at EOF with child-table attribution support)


# ΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉ
# MES DEMANDES ΓÇö LISTE COMBIN├ëE + STATISTIQUES
# ΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉΓòÉ

@frappe.whitelist()
def get_demandes_combined(limit=50, offset=0, type_filter=None, statut_filter=None):
    """
    Retourne la liste complète des demandes de l'utilisateur : 7 DocTypes.
    Filtrable par type (pse/pss/da/pv/pc/dc/bs) et par statut (workflow_state).
    """
    user = frappe.session.user
    if user == "Guest":
        return {"data": [], "total": 0}

    employee = frappe.db.get_value(
        "Employee",
        {"user_id": user, "status": "Active"},
        ["name", "employee_name"],
        as_dict=True,
    )
    emp_name = employee.name if employee else None

    results = []

    def _emp_or_owner_filter(statut=None):
        """Base filter pour les DocTypes avec champ employee."""
        f = {"employee": emp_name} if emp_name else {"owner": user}
        if statut:
            f["workflow_state"] = statut
        return f

    def _owner_filter(statut=None):
        f = {"owner": user}
        if statut:
            f["workflow_state"] = statut
        return f

    def _append(docs, type_label, type_code, route, date_field="creation", objet_field=None, montant_field=None):
        for doc in docs:
            date_val = getattr(doc, date_field, None) or doc.creation
            results.append({
                "name": doc.name,
                "type": type_label,
                "type_code": type_code,
                "objet": (getattr(doc, objet_field, "") if objet_field else "") or "",
                "date": str(date_val)[:10],
                "statut": doc.workflow_state or "Brouillon",
                "color": _STATE_COLORS.get(doc.workflow_state or "Brouillon", "gray"),
                "montant": getattr(doc, montant_field, 0) or 0 if montant_field else 0,
                "url": f"/{route}/{doc.name}",
                "creation": str(doc.creation),
            })

    # --- Permission Sortie Employé ---
    if not type_filter or type_filter == "pse":
        try:
            _append(frappe.get_all("Permission Sortie Employe",
                filters=_emp_or_owner_filter(statut_filter),
                fields=["name", "workflow_state", "date_sortie", "motif", "creation"],
                order_by="creation desc", limit_page_length=int(limit) * 4),
                "Permission Sortie", "pse", "permission-sortie-employe",
                date_field="date_sortie", objet_field="motif")
        except Exception:
            pass

    # --- Permission Sortie Stagiaire ---
    if not type_filter or type_filter == "pss":
        try:
            _append(frappe.get_all("Permission Sortie Stagiaire",
                filters=_emp_or_owner_filter(statut_filter),
                fields=["name", "workflow_state", "date_sortie", "motif", "creation"],
                order_by="creation desc", limit_page_length=int(limit) * 4),
                "Permission Stagiaire", "pss", "permission-sortie-stagiaire",
                date_field="date_sortie", objet_field="motif")
        except Exception:
            pass

    # --- Demande Achat ---
    if not type_filter or type_filter == "da":
        try:
            _append(frappe.get_all("Demande Achat KYA",
                filters=_emp_or_owner_filter(statut_filter),
                fields=["name", "workflow_state", "date_demande", "objet", "montant_total", "creation"],
                order_by="creation desc", limit_page_length=int(limit) * 4),
                "Demande d'Achat", "da", "demande-achat",
                date_field="date_demande", objet_field="objet", montant_field="montant_total")
        except Exception:
            pass

    # --- PV Sortie Matériel ---
    if not type_filter or type_filter == "pv":
        try:
            _append(frappe.get_all("PV Sortie Materiel",
                filters=_owner_filter(statut_filter),
                fields=["name", "workflow_state", "date_sortie", "objet", "creation"],
                order_by="creation desc", limit_page_length=int(limit) * 4),
                "PV Sortie Mat\u00e9riel", "pv", "pv-sortie-materiel",
                date_field="date_sortie", objet_field="objet")
        except Exception:
            pass

    # --- Planning Congé ---
    if not type_filter or type_filter == "pc":
        try:
            _append(frappe.get_all("Planning Conge",
                filters=_emp_or_owner_filter(statut_filter),
                fields=["name", "workflow_state", "annee", "creation"],
                order_by="creation desc", limit_page_length=int(limit) * 4),
                "Planning Cong\u00e9", "pc", "planning-conge",
                objet_field="annee")
        except Exception:
            pass

    # --- Demande Congé (Leave Application) ---
    if not type_filter or type_filter == "dc":
        try:
            _append(frappe.get_all("Leave Application",
                filters=_emp_or_owner_filter(statut_filter),
                fields=["name", "workflow_state", "from_date", "leave_type", "total_leave_days", "creation"],
                order_by="creation desc", limit_page_length=int(limit) * 4),
                "Cong\u00e9", "dc", "demande-conge",
                date_field="from_date", objet_field="leave_type")
        except Exception:
            pass

    # --- Bilan de Stage ---
    if not type_filter or type_filter == "bs":
        try:
            _append(frappe.get_all("Bilan de Stage",
                filters=_emp_or_owner_filter(statut_filter),
                fields=["name", "workflow_state", "date_debut", "creation"],
                order_by="creation desc", limit_page_length=int(limit) * 4),
                "Bilan de Stage", "bs", "bilan-fin-de-stage",
                date_field="date_debut")
        except Exception:
            pass

    # Tri global par date de création décroissante
    results.sort(key=lambda x: x["creation"], reverse=True)

    total = len(results)
    offset = int(offset)
    limit = int(limit)
    data = results[offset: offset + limit]

    return {"data": data, "total": total}


@frappe.whitelist()
def get_demandes_stats():
    """
    Retourne les statistiques des demandes de l'utilisateur courant :
    - total, par statut (brouillon / en_cours / approuve / rejete)
    - par type (7 DocTypes)
    """
    user = frappe.session.user
    if user == "Guest":
        return {}

    employee = frappe.db.get_value(
        "Employee",
        {"user_id": user, "status": "Active"},
        "name",
    )

    stats = {
        "total": 0,
        "brouillon": 0,
        "en_cours": 0,
        "approuve": 0,
        "rejete": 0,
        "pv_total": 0,
        "da_total": 0,
        "pse_total": 0,
        "pss_total": 0,
        "pc_total": 0,
        "dc_total": 0,
        "bs_total": 0,
    }

    en_cours_states = {
        "En attente Chef", "En attente Chef Service", "En attente RH",
        "En attente DG", "En attente DGA", "En attente Magasin",
        "En attente Audit", "En attente Comptabilit\u00e9",
        "En attente Resp. Stagiaires", "En cours",
        "En attente Approbation", "En attente DAAF",
        "En attente Direction", "En attente du Sup\u00e9rieur Imm\u00e9diat",
    }
    approuve_states = {"Approuv\u00e9", "Approuv\u00e9e", "Valid\u00e9", "Open"}
    rejete_states = {"Rejet\u00e9", "Rejet\u00e9e", "Annul\u00e9", "Rejected"}

    def _count(docs, stat_key):
        stats[stat_key] = len(docs)
        stats["total"] += len(docs)
        for doc in docs:
            ws = doc.workflow_state or "Brouillon"
            if ws == "Brouillon":
                stats["brouillon"] += 1
            elif ws in en_cours_states:
                stats["en_cours"] += 1
            elif ws in approuve_states:
                stats["approuve"] += 1
            elif ws in rejete_states:
                stats["rejete"] += 1

    emp_f = {"employee": employee} if employee else {"owner": user}

    try:
        _count(frappe.get_all("PV Sortie Materiel", filters={"owner": user},
                              fields=["workflow_state"]), "pv_total")
    except Exception:
        pass
    try:
        _count(frappe.get_all("Demande Achat KYA", filters=emp_f,
                              fields=["workflow_state"]), "da_total")
    except Exception:
        pass
    try:
        _count(frappe.get_all("Permission Sortie Employe", filters=emp_f,
                              fields=["workflow_state"]), "pse_total")
    except Exception:
        pass
    try:
        _count(frappe.get_all("Permission Sortie Stagiaire", filters=emp_f,
                              fields=["workflow_state"]), "pss_total")
    except Exception:
        pass
    try:
        _count(frappe.get_all("Planning Conge", filters=emp_f,
                              fields=["workflow_state"]), "pc_total")
    except Exception:
        pass
    try:
        _count(frappe.get_all("Leave Application", filters=emp_f,
                              fields=["workflow_state"]), "dc_total")
    except Exception:
        pass
    try:
        _count(frappe.get_all("Bilan de Stage", filters=emp_f,
                              fields=["workflow_state"]), "bs_total")
    except Exception:
        pass

    return stats


# ─── Dashboard Stagiaires ΓöÇΓöÇΓöÇΓöÇΓöÇΓöÇΓöÇΓöÇΓöÇΓöÇΓöÇΓöÇΓöÇΓöÇΓöÇΓöÇΓöÇΓöÇΓöÇΓöÇΓöÇΓöÇΓöÇΓöÇΓöÇΓöÇΓöÇΓöÇΓöÇΓöÇΓöÇΓöÇΓöÇΓöÇΓöÇΓöÇΓöÇΓöÇΓöÇΓöÇΓöÇΓöÇΓöÇΓöÇΓöÇΓöÇΓöÇΓöÇΓöÇΓöÇΓöÇΓöÇΓöÇΓöÇ

@frappe.whitelist()
def get_dashboard_stagiaires(annee=None):
    """Stats globales pour le module Stagiaires (accessible RH + System Manager)."""
    import datetime
    if annee:
        annee = int(annee)
    else:
        annee = datetime.date.today().year

    stats = {
        "annee": annee,
        "stagiaires_actifs": 0,
        "stagiaires_total": 0,
        "permissions_total": 0,
        "permissions_approuvees": 0,
        "permissions_en_cours": 0,
        "permissions_rejetees": 0,
        "bilans_total": 0,
        "bilans_soumis": 0,
        "note_moyenne": 0,
        "mentions": {},
        "permissions_par_mois": {},
        "top_beneficiaires": [],
    }

    # Stagiaires
    try:
        stagiaires = frappe.get_all(
            "Employee",
            filters={"employment_type": "Stage", "status": "Active"},
            fields=["name", "employee_name", "department"],
        )
        stats["stagiaires_actifs"] = len(stagiaires)
        stats["stagiaires_total"] = frappe.db.count(
            "Employee", {"employment_type": "Stage"}
        )
    except Exception:
        pass

    en_cours_states = {
        "En attente Chef", "En attente Chef Service",
        "En attente Resp. Stagiaires", "En attente DG", "En cours",
    }
    approuve_states = {"Approuvé", "Approuvée", "Validé"}
    rejete_states = {"Rejeté", "Rejetée", "Annulé"}

    # Permissions Stagiaires
    try:
        perms = frappe.get_all(
            "Permission Sortie Stagiaire",
            filters=[["date_sortie", ">=", f"{annee}-01-01"],
                     ["date_sortie", "<=", f"{annee}-12-31"]],
            fields=["name", "workflow_state", "date_sortie", "employee", "employee_name"],
        )
        stats["permissions_total"] = len(perms)
        mois_count = {}
        benef_count = {}
        for p in perms:
            ws = p.workflow_state or "Brouillon"
            if ws in approuve_states:
                stats["permissions_approuvees"] += 1
            elif ws in en_cours_states:
                stats["permissions_en_cours"] += 1
            elif ws in rejete_states:
                stats["permissions_rejetees"] += 1
            # Par mois
            if p.date_sortie:
                mois = str(p.date_sortie)[:7]
                mois_count[mois] = mois_count.get(mois, 0) + 1
            # Par b├⌐n├⌐ficiaire
            emp = p.employee_name or p.employee or "Inconnu"
            benef_count[emp] = benef_count.get(emp, 0) + 1
        stats["permissions_par_mois"] = mois_count
        # Top 5
        stats["top_beneficiaires"] = sorted(
            [{"nom": k, "count": v} for k, v in benef_count.items()],
            key=lambda x: x["count"], reverse=True
        )[:5]
    except Exception:
        pass

    # Bilans de Stage
    try:
        bilans = frappe.get_all(
            "Bilan Fin de Stage",
            fields=["name", "workflow_state", "note_globale", "mention"],
        )
        stats["bilans_total"] = len(bilans)
        notes = []
        mentions = {}
        for b in bilans:
            ws = b.workflow_state or "Brouillon"
            if ws in approuve_states:
                stats["bilans_soumis"] += 1
            if b.note_globale:
                notes.append(float(b.note_globale))
            if b.mention:
                mentions[b.mention] = mentions.get(b.mention, 0) + 1
        if notes:
            stats["note_moyenne"] = round(sum(notes) / len(notes), 2)
        stats["mentions"] = mentions
    except Exception:
        pass

    return stats


@frappe.whitelist()
def get_dashboard_employes(annee=None):
    """Stats globales pour le module Employ├⌐s (accessible RH + System Manager)."""
    import datetime
    if annee:
        annee = int(annee)
    else:
        annee = datetime.date.today().year

    stats = {
        "annee": annee,
        "employes_actifs": 0,
        "employes_total": 0,
        "permissions_total": 0,
        "permissions_approuvees": 0,
        "permissions_en_cours": 0,
        "permissions_rejetees": 0,
        "permissions_par_mois": {},
        "conges_total": 0,
        "conges_approuves": 0,
        "pv_total": 0,
        "pv_approuves": 0,
        "da_total": 0,
        "da_approuves": 0,
        "da_montant_total_approuve": 0,
        "absences_total": 0,
        "presences_total": 0,
        "taux_presence": 0,
        "top_demandeurs_da": [],
    }

    en_cours_states = {
        "En attente Chef", "En attente Chef Service", "En attente RH",
        "En attente DG", "En attente DGA", "En attente Magasin",
        "En attente Audit", "En attente Comptabilité",
        "En attente Approbation", "En cours",
    }
    approuve_states = {"Approuvé", "Approuvée", "Validé"}
    rejete_states = {"Rejeté", "Rejetée", "Annulé"}

    # Employ├⌐s (hors stagiaires)
    try:
        emp_filters = [["employment_type", "!=", "Stage"], ["status", "=", "Active"]]
        stats["employes_actifs"] = frappe.db.count("Employee", emp_filters)
        stats["employes_total"] = frappe.db.count(
            "Employee", [["employment_type", "!=", "Stage"]]
        )
    except Exception:
        pass

    # Permissions Employ├⌐s
    try:
        perms = frappe.get_all(
            "Permission Sortie Employe",
            filters=[["date_sortie", ">=", f"{annee}-01-01"],
                     ["date_sortie", "<=", f"{annee}-12-31"]],
            fields=["name", "workflow_state", "date_sortie"],
        )
        stats["permissions_total"] = len(perms)
        mois_count = {}
        for p in perms:
            ws = p.workflow_state or "Brouillon"
            if ws in approuve_states:
                stats["permissions_approuvees"] += 1
            elif ws in en_cours_states:
                stats["permissions_en_cours"] += 1
            elif ws in rejete_states:
                stats["permissions_rejetees"] += 1
            if p.date_sortie:
                mois = str(p.date_sortie)[:7]
                mois_count[mois] = mois_count.get(mois, 0) + 1
        stats["permissions_par_mois"] = mois_count
    except Exception:
        pass

    # Plannings Cong├⌐
    try:
        conges = frappe.get_all(
            "Planning Conge",
            filters=[["annee", "=", annee]],
            fields=["name", "workflow_state"],
        )
        stats["conges_total"] = len(conges)
        for c in conges:
            if (c.workflow_state or "") in approuve_states:
                stats["conges_approuves"] += 1
    except Exception:
        pass

    # PV Sortie Mat├⌐riel
    try:
        pvs = frappe.get_all(
            "PV Sortie Materiel",
            filters=[["date_sortie", ">=", f"{annee}-01-01"],
                     ["date_sortie", "<=", f"{annee}-12-31"]],
            fields=["name", "workflow_state"],
        )
        stats["pv_total"] = len(pvs)
        for p in pvs:
            if (p.workflow_state or "") in approuve_states:
                stats["pv_approuves"] += 1
    except Exception:
        pass

    # Demandes d'Achat
    try:
        das = frappe.get_all(
            "Demande Achat KYA",
            filters=[["date_demande", ">=", f"{annee}-01-01"],
                     ["date_demande", "<=", f"{annee}-12-31"]],
            fields=["name", "workflow_state", "montant_total", "employee"],
        )
        stats["da_total"] = len(das)
        da_by_emp = {}
        for d in das:
            ws = d.workflow_state or "Brouillon"
            if ws in approuve_states:
                stats["da_approuves"] += 1
                stats["da_montant_total_approuve"] += float(d.montant_total or 0)
            emp = d.employee or "Inconnu"
            da_by_emp[emp] = da_by_emp.get(emp, {"count": 0, "montant": 0})
            da_by_emp[emp]["count"] += 1
            da_by_emp[emp]["montant"] += float(d.montant_total or 0)
        stats["top_demandeurs_da"] = sorted(
            [{"emp": k, "count": v["count"], "montant": v["montant"]}
             for k, v in da_by_emp.items()],
            key=lambda x: x["count"], reverse=True
        )[:5]
    except Exception:
        pass

    # Pr├⌐sences / Absences
    try:
        att = frappe.get_all(
            "Attendance",
            filters=[["attendance_date", ">=", f"{annee}-01-01"],
                     ["attendance_date", "<=", f"{annee}-12-31"]]
            if annee else [],
            fields=["status"],
        )
        for a in att:
            if a.status in ("Present", "Work From Home", "Half Day"):
                stats["presences_total"] += 1
            elif a.status == "Absent":
                stats["absences_total"] += 1
        total_att = stats["presences_total"] + stats["absences_total"]
        if total_att > 0:
            stats["taux_presence"] = round(
                stats["presences_total"] / total_att * 100, 1
            )
    except Exception:
        pass

    return stats


@frappe.whitelist()
def get_webforms_list():
    """Retourne la liste des web forms KYA HR avec leurs URLs et statuts."""
    forms = [
        {
            "name": "permission-sortie-stagiaire",
            "label": "Permission de Sortie Stagiaire",
            "route": "/permission-sortie-stagiaire",
            "new_route": "/permission-sortie-stagiaire/new",
            "ref": "AEA-ENG-30-V01",
            "module": "Stagiaires",
            "icon": "≡ƒÄô"
        },
        {
            "name": "permission-sortie-employe",
            "label": "Permission de Sortie Employ├⌐",
            "route": "/permission-sortie-employe",
            "new_route": "/permission-sortie-employe/new",
            "ref": "AEA-ENG-30-V01",
            "module": "Employ├⌐s",
            "icon": "≡ƒæñ"
        },
        {
            "name": "demande-achat",
            "label": "Demande d'Achat",
            "route": "/demande-achat",
            "new_route": "/demande-achat/new",
            "ref": "AEA-ENG-30-V01",
            "module": "Achats",
            "icon": "≡ƒ¢Æ"
        },
        {
            "name": "pv-sortie-materiel",
            "label": "PV Sortie de Mat├⌐riel",
            "route": "/pv-sortie-materiel",
            "new_route": "/pv-sortie-materiel/new",
            "ref": "AEA-ENG-30-V01",
            "module": "Stock",
            "icon": "≡ƒôª"
        },
        {
            "name": "planning-conge",
            "label": "Planning de Cong├⌐",
            "route": "/planning-conge",
            "new_route": "/planning-conge/new",
            "ref": "AEA-ENG-30-V01",
            "module": "Employ├⌐s",
            "icon": "≡ƒôà"
        },
        {
            "name": "bilan-fin-de-stage",
            "label": "Bilan de Fin de Stage",
            "route": "/bilan-fin-de-stage",
            "new_route": "/bilan-fin-de-stage/new",
            "ref": "AEA-ENG-30-V01",
            "module": "Stagiaires",
            "icon": "≡ƒôï"
        },
    ]
    # Enrichir avec le count des documents li├⌐s
    doctypes_map = {
        "permission-sortie-stagiaire": "Permission Sortie Stagiaire",
        "permission-sortie-employe": "Permission Sortie Employe",
        "demande-achat": "Demande Achat KYA",
        "pv-sortie-materiel": "PV Sortie Materiel",
        "planning-conge": "Planning Conge",
        "bilan-fin-de-stage": "Bilan Fin de Stage",
    }
    for form in forms:
        dt = doctypes_map.get(form["name"])
        if dt:
            try:
                form["count"] = frappe.db.count(dt)
            except Exception:
                form["count"] = 0
    return forms


# ═══════════════════════════════════════════════════════════════════════════════
#  INDICATEURS DE PERFORMANCE (KYA Indicator)
# ═══════════════════════════════════════════════════════════════════════════════

def _quarter_date_range(trimestre, annee):
    """Return (from_date, to_date) strings for a given quarter."""
    mapping = {
        "T1": ("{}-01-01", "{}-03-31"),
        "T2": ("{}-04-01", "{}-06-30"),
        "T3": ("{}-07-01", "{}-09-30"),
        "T4": ("{}-10-01", "{}-12-31"),
    }
    tpl = mapping.get(trimestre, ("{}-01-01", "{}-12-31"))
    return tpl[0].format(annee), tpl[1].format(annee)


def _calc_presenteisme(employee, from_date, to_date):
    """Calculate presenteeism rate from Attendance records."""
    try:
        rows = frappe.db.sql(
            """SELECT status FROM tabAttendance
               WHERE employee=%s AND attendance_date BETWEEN %s AND %s
               AND docstatus=1""",
            (employee, from_date, to_date), as_dict=True,
        )
        if not rows:
            return None
        present = sum(1 for r in rows if r.status in ("Present", "Work From Home", "Half Day"))
        total = len(rows)
        return round(present / total * 100, 2) if total else None
    except Exception:
        return None


def _calc_perf_pro(employee, from_date, to_date):
    """Calcule la perf professionnelle depuis Tache Equipe.
    = moyenne des taux_effectif des tâches où l'employé est attributaire."""
    try:
        if not frappe.db.table_exists("tabTache Equipe"):
            return None
        taches = frappe.db.sql(
            """SELECT te.taux_effectif
               FROM `tabTache Equipe` te
               JOIN `tabTache Equipe Attribution` tea ON tea.parent = te.name
               JOIN `tabPlan Trimestriel` pt ON pt.name = te.plan
               WHERE tea.employe = %s
                 AND CONCAT(pt.trimestre, ' ', pt.annee) BETWEEN %s AND %s""",
            (employee,
             from_date[:7],   # Approx: use year range fallback below
             to_date[:7]),
            as_dict=True,
        )
        # Fallback: join via plan trimestre/annee (more robust)
        taches = frappe.db.sql(
            """SELECT te.taux_effectif
               FROM `tabTache Equipe` te
               JOIN `tabTache Equipe Attribution` tea ON tea.parent = te.name
               JOIN `tabPlan Trimestriel` pt ON pt.name = te.plan
               WHERE tea.employe = %s
                 AND te.creation BETWEEN %s AND %s""",
            (employee, from_date, to_date),
            as_dict=True,
        )
        if not taches:
            return None
        vals = [float(t.taux_effectif) for t in taches if t.taux_effectif is not None]
        return round(sum(vals) / len(vals), 2) if vals else None
    except Exception:
        return None


def _calc_sat_evaluations(employee, trimestre, annee, eval_type):
    """Calcule le taux moyen d'évaluation pour un type donné.
    eval_type: 'N+1 évalue N' ou 'N évalue N+1'."""
    try:
        if not frappe.db.table_exists("tabKYA Evaluation"):
            return None
        if eval_type == "N+1 \u00e9value N":
            # N+1 evaluates this employee → employee is "evalue"
            evals = frappe.db.sql(
                """SELECT taux_moyen FROM `tabKYA Evaluation`
                   WHERE evalue=%s AND trimestre=%s AND annee=%s
                     AND type_evaluation=%s AND statut IN ('Soumis','Valid\u00e9')""",
                (employee, trimestre, annee, eval_type), as_dict=True,
            )
        else:
            # This employee evaluates their N+1 → employee is "evaluateur"
            evals = frappe.db.sql(
                """SELECT taux_moyen FROM `tabKYA Evaluation`
                   WHERE evaluateur=%s AND trimestre=%s AND annee=%s
                     AND type_evaluation=%s AND statut IN ('Soumis','Valid\u00e9')""",
                (employee, trimestre, annee, eval_type), as_dict=True,
            )
        if not evals:
            return None
        vals = [float(e.taux_moyen) for e in evals if e.taux_moyen is not None]
        return round(sum(vals) / len(vals), 2) if vals else None
    except Exception:
        return None


def _calc_sat_personnel(department, from_date, to_date):
    """Calcule la satisfaction personnel depuis les réponses KYA Form (enquêtes de satisfaction)."""
    try:
        if not frappe.db.table_exists("tabKYA Form Response"):
            return None
        # Look for KYA Form Responses linked to satisfaction survey forms for this department
        # The KYA Form should have a tag/category "Satisfaction Personnel"
        # Count satisfaction responses (score_global is calculated externally by admin)
        # Use a simple count proxy: 100 if at least 1 satisfaction form response, else None
        count = frappe.db.sql(
            """SELECT COUNT(*) as cnt
               FROM `tabKYA Form Response` kfr
               JOIN `tabKYA Form` kf ON kf.name = kfr.formulaire
               WHERE kf.titre LIKE %s
                 AND kfr.soumis_le BETWEEN %s AND %s
                 AND kfr.soumis_le IS NOT NULL""",
            ("%Satisfaction%", from_date, to_date), as_dict=True,
        )
        n = count[0].cnt if count else 0
        if n == 0:
            return None
        # Calculer la moyenne des réponses numériques pour les enquêtes de satisfaction
        ans_rows = frappe.db.sql(
            """SELECT kfa.valeur
               FROM `tabKYA Form Answer` kfa
               JOIN `tabKYA Form Response` kfr ON kfr.name = kfa.parent
               JOIN `tabKYA Form` kf ON kf.name = kfr.formulaire
               WHERE kf.titre LIKE %s
                 AND kfr.soumis_le BETWEEN %s AND %s
                 AND kfa.valeur REGEXP '^[0-9]+\\.?[0-9]*$'""",
            ("%Satisfaction%", from_date, to_date), as_dict=True,
        )
        vals = []
        for r in ans_rows:
            try:
                vals.append(float(r.valeur))
            except (ValueError, TypeError):
                pass
        return round(sum(vals) / len(vals), 2) if vals else None
    except Exception:
        return None


def _calc_non_accident(employee, from_date, to_date):
    """Vérifie si l'employé a eu un accident de travail sur la période.
    Retourne 100 (taux non-accident) si aucun, 0 si accident détecté."""
    try:
        # Check Leave Application with leave_type matching "Accident"
        count = frappe.db.count(
            "Leave Application",
            filters=[
                ["employee", "=", employee],
                ["from_date", ">=", from_date],
                ["to_date", "<=", to_date],
                ["leave_type", "like", "%Accident%"],
                ["docstatus", "=", 1],
            ],
        )
        return 0.0 if count > 0 else 100.0
    except Exception:
        return None


def _calc_redressement(employee, from_date, to_date):
    """Vérifie s'il y a eu une mesure disciplinaire (Warning/Corrective Action) sur la période.
    Retourne 1 si mesure disciplinaire, 0 sinon."""
    try:
        # Check Employee Warning (HRMS) or HR Notice if available
        for dt in ("Employee Warning", "Employee Grievance"):
            if frappe.db.table_exists("tab" + dt):
                count = frappe.db.count(
                    dt,
                    filters=[
                        ["employee", "=", employee],
                        ["creation", ">=", from_date],
                        ["creation", "<=", to_date],
                        ["docstatus", "=", 1],
                    ],
                )
                if count > 0:
                    return 1.0
        return 0.0
    except Exception:
        return None


@frappe.whitelist()
def get_employee_indicators(employee, trimestre=None, annee=None):
    """Calcule ou récupère les indicateurs KPI pour un employé et un trimestre donné.
    Si trimestre/annee non fournis, utilise le trimestre en cours.

    Retourne:
    {
        employee, employee_name, department, trimestre, annee,
        presenteisme, perf_pro, sat_n_1, sat_n_plus_1,
        sat_clients, sat_personnel, non_accident, redressement,
        from_db  // True si lu depuis KYA Indicator, False si calculé à la volée
    }
    """
    import datetime

    today = datetime.date.today()
    if not annee:
        annee = today.year
    annee = int(annee)
    if not trimestre:
        m = today.month
        trimestre = "T1" if m <= 3 else "T2" if m <= 6 else "T3" if m <= 9 else "T4"

    # Check permissions: HR roles + Chef d'Équipe can see all, Employee sees own data only
    user = frappe.session.user
    user_roles = set(frappe.get_roles(user))
    hr_roles = {"System Manager", "HR Manager", "HR User",
                "Accounts Manager", "Purchase Manager"}
    if not (user_roles & hr_roles):
        # Non-HR: can only see own indicators
        own_emp = frappe.db.get_value("Employee", {"user_id": user, "status": "Active"}, "name")
        if own_emp != employee:
            # Chef d'Équipe can see their team
            is_chef = frappe.db.exists("Employee", {"name": employee, "reports_to": own_emp})
            if not is_chef:
                frappe.throw("Acc\u00e8s refus\u00e9", frappe.PermissionError)

    emp_doc = frappe.db.get_value(
        "Employee", employee,
        ["employee_name", "department", "designation"], as_dict=True,
    )
    if not emp_doc:
        frappe.throw(f"Employ\u00e9 introuvable: {employee}")

    result = {
        "employee": employee,
        "employee_name": emp_doc.employee_name,
        "department": emp_doc.department,
        "designation": emp_doc.designation,
        "trimestre": trimestre,
        "annee": annee,
        "presenteisme": None,
        "perf_pro": None,
        "sat_n_1": None,
        "sat_n_plus_1": None,
        "sat_clients": None,
        "sat_personnel": None,
        "non_accident": None,
        "redressement": None,
        "from_db": False,
    }

    # 1. Try to load pre-calculated from KYA Indicator
    if frappe.db.table_exists("tabKYA Indicator"):
        saved = frappe.db.get_value(
            "KYA Indicator",
            {"employee": employee, "trimestre": trimestre, "annee": annee},
            ["presenteisme", "perf_pro", "sat_n_1", "sat_n_plus_1",
             "sat_clients", "sat_personnel", "non_accident", "redressement"],
            as_dict=True,
        )
        if saved:
            result.update({
                "presenteisme":  saved.presenteisme,
                "perf_pro":      saved.perf_pro,
                "sat_n_1":       saved.sat_n_1,
                "sat_n_plus_1":  saved.sat_n_plus_1,
                "sat_clients":   saved.sat_clients,
                "sat_personnel": saved.sat_personnel,
                "non_accident":  saved.non_accident,
                "redressement":  saved.redressement,
                "from_db":       True,
            })
            return result

    # 2. Calculate on the fly
    from_date, to_date = _quarter_date_range(trimestre, annee)
    result["presenteisme"]  = _calc_presenteisme(employee, from_date, to_date)
    result["perf_pro"]      = _calc_perf_pro(employee, from_date, to_date)
    result["sat_n_1"]       = _calc_sat_evaluations(employee, trimestre, annee, "N+1 \u00e9value N")
    result["sat_n_plus_1"]  = _calc_sat_evaluations(employee, trimestre, annee, "N \u00e9value N+1")
    result["sat_clients"]   = None   # Manual — not calculated
    result["sat_personnel"] = _calc_sat_personnel(emp_doc.department, from_date, to_date)
    result["non_accident"]  = _calc_non_accident(employee, from_date, to_date)
    result["redressement"]  = _calc_redressement(employee, from_date, to_date)
    return result


@frappe.whitelist()
def get_team_indicators(department=None, trimestre=None, annee=None):
    """Retourne les indicateurs agrégés pour toute une équipe / département."""
    import datetime

    today = datetime.date.today()
    if not annee:
        annee = today.year
    annee = int(annee)
    if not trimestre:
        m = today.month
        trimestre = "T1" if m <= 3 else "T2" if m <= 6 else "T3" if m <= 9 else "T4"

    allowed_roles = {"System Manager", "HR Manager", "HR User",
                     "Accounts Manager", "Purchase Manager"}
    user_roles = set(frappe.get_roles(frappe.session.user))
    if not (user_roles & allowed_roles):
        # Chef d'Équipe sees only their own department
        own_emp = frappe.db.get_value(
            "Employee", {"user_id": frappe.session.user, "status": "Active"}, "name"
        )
        if own_emp:
            department = frappe.db.get_value("Employee", own_emp, "department")
        if not department:
            frappe.throw("Acc\u00e8s refus\u00e9", frappe.PermissionError)

    filters = [["status", "=", "Active"], ["employment_type", "!=", "Stage"]]
    if department:
        filters.append(["department", "=", department])

    employees = frappe.get_all("Employee", filters=filters,
                               fields=["name", "employee_name", "department", "designation"])

    results = []
    for emp in employees:
        ind = get_employee_indicators(emp.name, trimestre, annee)
        results.append(ind)
    return {"employees": results, "trimestre": trimestre, "annee": annee}


@frappe.whitelist()
def save_employee_indicators(employee, trimestre, annee, sat_clients=None, notes=None):
    """Calcule et sauvegarde les indicateurs dans KYA Indicator DocType.
    Réservé aux rôles RH. sat_clients saisie manuelle."""
    allowed_roles = {"System Manager", "HR Manager"}
    user_roles = set(frappe.get_roles(frappe.session.user))
    if not (user_roles & allowed_roles):
        frappe.throw("Acc\u00e8s r\u00e9serv\u00e9 aux managers RH", frappe.PermissionError)

    if not frappe.db.table_exists("tabKYA Indicator"):
        frappe.throw("DocType KYA Indicator non install\u00e9 — ex\u00e9cutez bench migrate")

    # Calculate live
    from_date, to_date = _quarter_date_range(trimestre, int(annee))
    emp_doc = frappe.db.get_value("Employee", employee,
                                  ["employee_name", "department"], as_dict=True)

    values = {
        "presenteisme":  _calc_presenteisme(employee, from_date, to_date),
        "perf_pro":      _calc_perf_pro(employee, from_date, to_date),
        "sat_n_1":       _calc_sat_evaluations(employee, trimestre, annee, "N+1 \u00e9value N"),
        "sat_n_plus_1":  _calc_sat_evaluations(employee, trimestre, annee, "N \u00e9value N+1"),
        "sat_clients":   float(sat_clients) if sat_clients is not None else None,
        "sat_personnel": _calc_sat_personnel(emp_doc.department, from_date, to_date),
        "non_accident":  _calc_non_accident(employee, from_date, to_date),
        "redressement":  _calc_redressement(employee, from_date, to_date),
    }

    existing = frappe.db.get_value(
        "KYA Indicator",
        {"employee": employee, "trimestre": trimestre, "annee": annee},
        "name",
    )
    if existing:
        doc = frappe.get_doc("KYA Indicator", existing)
        for k, v in values.items():
            if v is not None:
                setattr(doc, k, v)
        if notes:
            doc.notes = notes
        doc.save(ignore_permissions=True)
    else:
        doc = frappe.new_doc("KYA Indicator")
        doc.employee  = employee
        doc.trimestre = trimestre
        doc.annee     = int(annee)
        doc.notes     = notes or ""
        for k, v in values.items():
            setattr(doc, k, v)
        doc.insert(ignore_permissions=True)

    frappe.db.commit()
    return {"name": doc.name, "values": values}


# ═══════════════════════════════════════════════════════════════════════════════
#  GESTION ÉQUIPE — CONTEXTE CHEF D'ÉQUIPE (Isolation département)
# ═══════════════════════════════════════════════════════════════════════════════

@frappe.whitelist()
def get_chef_equipe_context():
    """Retourne le contexte département du chef d'équipe connecté.

    Les chefs d'équipe (rôle Chef Service) ne voient QUE leur département.
    Les rôles HR/Manager voient tout → department=None (pas de filtre).

    Retourne :
    {
      employee, employee_name, department, department_name,
      is_hr_role  (bool),  → True si HR/Manager (accès total)
      team_members  → liste des noms d'employés du département
    }
    """
    user = frappe.session.user
    if user == "Guest":
        frappe.throw("Authentification requise", frappe.PermissionError)

    hr_roles = {"System Manager", "HR Manager", "HR User",
                "Accounts Manager", "Purchase Manager"}
    user_roles = set(frappe.get_roles(user))
    is_hr = bool(user_roles & hr_roles)

    emp = frappe.db.get_value(
        "Employee",
        {"user_id": user, "status": "Active"},
        ["name", "employee_name", "department"],
        as_dict=True,
    )

    department = None
    department_name = None
    team_members = []

    if not is_hr and emp:
        department = emp.department
        if department:
            department_name = frappe.db.get_value("Department", department, "department_name") or department
            team_members = frappe.db.sql_list(
                "SELECT name FROM `tabEmployee` WHERE department=%s AND status='Active'",
                department,
            )

    return {
        "employee": emp.name if emp else None,
        "employee_name": emp.employee_name if emp else None,
        "department": department,
        "department_name": department_name,
        "is_hr_role": is_hr,
        "team_members": team_members,
    }


@frappe.whitelist()
def get_team_attendance_dashboard(date=None, department=None):
    """Version sécurisée de get_attendance_dashboard avec isolation département.

    - Rôles HR : peuvent passer n'importe quel department (ou None pour tout)
    - Rôle Chef Service : filtre automatiquement sur son département
    - Autres employés : voient uniquement leur propre fiche
    """
    user = frappe.session.user
    if user == "Guest":
        frappe.throw("Authentification requise", frappe.PermissionError)

    hr_roles = {"System Manager", "HR Manager", "HR User", "Accounts Manager"}
    user_roles = set(frappe.get_roles(user))
    is_hr = bool(user_roles & hr_roles)

    emp = frappe.db.get_value(
        "Employee",
        {"user_id": user, "status": "Active"},
        ["name", "department"],
        as_dict=True,
    )

    if not is_hr and emp:
        # Chef d'équipe → forcer son département
        department = emp.department

    return get_attendance_dashboard(date=date, department=department)


@frappe.whitelist()
def get_team_plans(department=None, trimestre=None, annee=None):
    """Plans Trimestriels filtrés par département avec isolation chef d'équipe.

    - Rôles HR : voient tous les plans (department=None = tous)
    - Chef Service : voient uniquement leur département
    """
    import datetime

    user = frappe.session.user
    if user == "Guest":
        frappe.throw("Authentification requise", frappe.PermissionError)

    hr_roles = {"System Manager", "HR Manager", "HR User"}
    user_roles = set(frappe.get_roles(user))
    is_hr = bool(user_roles & hr_roles)

    if not is_hr:
        emp = frappe.db.get_value(
            "Employee",
            {"user_id": user, "status": "Active"},
            ["name", "department"],
            as_dict=True,
        )
        if emp and emp.department:
            department = emp.department
        else:
            return {"plans": [], "department": None, "is_hr_role": False}

    if not frappe.db.table_exists("tabPlan Trimestriel"):
        return {"plans": [], "department": department, "is_hr_role": is_hr}

    today = datetime.date.today()
    if not annee:
        annee = today.year
    annee = int(annee)
    if not trimestre:
        m = today.month
        trimestre = "T1" if m <= 3 else "T2" if m <= 6 else "T3" if m <= 9 else "T4"

    filters = [["annee", "=", annee], ["trimestre", "=", trimestre]]
    if department:
        filters.append(["equipe", "=", department])

    plans = frappe.get_all(
        "Plan Trimestriel",
        filters=filters,
        fields=["name", "titre", "equipe", "trimestre", "annee",
                "statut", "chef_equipe", "progression_globale"],
        order_by="creation desc",
    )

    # Enrichir avec le nombre de tâches
    for p in plans:
        if frappe.db.table_exists("tabTache Equipe"):
            p["nb_taches"] = frappe.db.count("Tache Equipe", {"plan": p.name})
            p["nb_terminees"] = frappe.db.count(
                "Tache Equipe", {"plan": p.name, "statut": ["in", ["Terminé", "Validé"]]}
            )
        else:
            p["nb_taches"] = 0
            p["nb_terminees"] = 0
        p["url"] = f"/app/plan-trimestriel/{p.name}"

    return {
        "plans": plans,
        "department": department,
        "trimestre": trimestre,
        "annee": annee,
        "is_hr_role": is_hr,
    }


# ═══════════════════════════════════════════════════════════════════════════════
#  IMPORT PRÉSENCES DEPUIS EXCEL (Espace RH)
# ═══════════════════════════════════════════════════════════════════════════════

@frappe.whitelist()
def import_attendance_excel(file_url, employee_col="Matricule", date_col="Date",
                            time_col="Heure", log_type_col=None,
                            create_missing_employees=0):
    """Importe des présences/pointages depuis un fichier Excel.

    Paramètres (colonnes Excel, noms ou index lettre ex: "A", "B"):
      file_url : URL Frappe du fichier (ex: /files/pointage.xlsx)
      employee_col : colonne contenant le matricule ou ID employé (défaut "Matricule")
      date_col     : colonne date (défaut "Date", format YYYY-MM-DD ou DD/MM/YYYY)
      time_col     : colonne heure (défaut "Heure", format HH:MM ou HH:MM:SS)
      log_type_col : colonne IN/OUT (optionnelle)
      create_missing_employees : 0/1 (créer un checkin même si l'employé est introuvable par ID)

    Retourne :
      { imported, skipped, errors, checkins_created }
    """
    import re

    allowed_roles = {"System Manager", "HR Manager", "HR User"}
    user_roles = set(frappe.get_roles(frappe.session.user))
    if not (user_roles & allowed_roles):
        frappe.throw("Accès réservé au personnel RH", frappe.PermissionError)

    # Récupérer le fichier
    try:
        from frappe.utils.file_manager import get_file_path
        file_path = get_file_path(file_url)
    except Exception as e:
        frappe.throw(f"Fichier introuvable : {file_url} — {e}")

    # Lire le fichier Excel avec openpyxl
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active
    except ImportError:
        frappe.throw("Module openpyxl non disponible. Installez-le : pip install openpyxl")
    except Exception as e:
        frappe.throw(f"Impossible de lire le fichier Excel : {e}")

    # Lire l'en-tête (ligne 1)
    headers = []
    for cell in next(ws.iter_rows(max_row=1, values_only=True)):
        headers.append(str(cell).strip() if cell is not None else "")

    def _col_idx(col_name):
        """Retourne l'index (0-based) de la colonne par nom d'en-tête ou lettre Excel."""
        if isinstance(col_name, int):
            return col_name
        # Essai par nom d'en-tête
        col_name_lower = col_name.strip().lower()
        for i, h in enumerate(headers):
            if h.lower() == col_name_lower:
                return i
        # Essai par lettre Excel (A=0, B=1…)
        if re.match(r'^[A-Za-z]+$', col_name.strip()):
            idx = 0
            for c in col_name.strip().upper():
                idx = idx * 26 + (ord(c) - ord('A') + 1)
            return idx - 1
        raise ValueError(f"Colonne '{col_name}' non trouvée (en-têtes : {headers})")

    try:
        idx_emp  = _col_idx(employee_col)
        idx_date = _col_idx(date_col)
        idx_time = _col_idx(time_col)
        idx_log  = _col_idx(log_type_col) if log_type_col else None
    except ValueError as e:
        frappe.throw(str(e))

    # Construire le mapping matricule → Employee.name
    emp_map = {}
    for row in frappe.db.sql(
        "SELECT name, attendance_device_id FROM `tabEmployee` WHERE status='Active'",
        as_dict=True,
    ):
        if row.attendance_device_id:
            emp_map[str(row.attendance_device_id).strip()] = row.name
        emp_map[str(row.name).strip()] = row.name

    imported = 0
    skipped = 0
    errors = []
    checkins_created = []

    from frappe.utils import get_datetime
    import datetime as dt_mod

    def _parse_date(val):
        if isinstance(val, (dt_mod.date, dt_mod.datetime)):
            return val.strftime("%Y-%m-%d")
        s = str(val).strip()
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
            try:
                return dt_mod.datetime.strptime(s, fmt).strftime("%Y-%m-%d")
            except ValueError:
                continue
        raise ValueError(f"Format date inconnu : {s}")

    def _parse_time(val):
        if isinstance(val, dt_mod.time):
            return val.strftime("%H:%M:%S")
        if isinstance(val, (dt_mod.datetime,)):
            return val.strftime("%H:%M:%S")
        s = str(val).strip()
        # HH:MM or HH:MM:SS
        m = re.match(r'^(\d{1,2}):(\d{2})(?::(\d{2}))?$', s)
        if m:
            h, mi, sec = m.group(1), m.group(2), m.group(3) or "00"
            return f"{int(h):02d}:{mi}:{sec}"
        raise ValueError(f"Format heure inconnu : {s}")

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or all(v is None for v in row):
            continue
        try:
            raw_emp  = row[idx_emp]
            raw_date = row[idx_date]
            raw_time = row[idx_time]
            log_type = "IN"
            if idx_log is not None:
                lt_val = str(row[idx_log] or "").strip().upper()
                log_type = "OUT" if lt_val in ("OUT", "S", "SORTIE", "0") else "IN"

            if raw_emp is None or raw_date is None or raw_time is None:
                skipped += 1
                continue

            matricule = str(raw_emp).strip()
            date_str  = _parse_date(raw_date)
            time_str  = _parse_time(raw_time)

            employee_id = emp_map.get(matricule)
            if not employee_id:
                errors.append(f"Ligne {row_num}: matricule '{matricule}' introuvable")
                skipped += 1
                continue

            # Vérifier doublon
            ts_str = f"{date_str} {time_str}"
            existing = frappe.db.exists("Employee Checkin", {
                "employee": employee_id,
                "time": ["between", [f"{date_str} {time_str[:-2]}00", f"{date_str} {time_str[:-2]}59"]],
                "log_type": log_type,
            })
            if existing:
                skipped += 1
                continue

            checkin = frappe.get_doc({
                "doctype": "Employee Checkin",
                "employee": employee_id,
                "time": ts_str,
                "log_type": log_type,
                "device_id": "import-excel",
            })
            checkin.insert(ignore_permissions=True)
            imported += 1
            checkins_created.append({
                "employee": employee_id,
                "time": ts_str,
                "log_type": log_type,
            })
        except Exception as e:
            errors.append(f"Ligne {row_num}: {e}")
            skipped += 1

    frappe.db.commit()
    return {
        "imported": imported,
        "skipped": skipped,
        "errors": errors[:50],  # Limiter à 50 erreurs
        "checkins_created": checkins_created[:20],  # Aperçu des 20 premiers
        "total_rows": row_num - 1 if 'row_num' in dir() else 0,
    }


@frappe.whitelist()
def get_rh_attendance_import_info():
    """Retourne les informations de mapping pour l'import Excel de présences.
    Aide le RH à comprendre quels champs sont attendus."""
    allowed_roles = {"System Manager", "HR Manager", "HR User"}
    user_roles = set(frappe.get_roles(frappe.session.user))
    if not (user_roles & allowed_roles):
        frappe.throw("Accès réservé au personnel RH", frappe.PermissionError)

    # Employés actifs avec leur device_id pour l'aide au mapping
    employees = frappe.db.sql(
        """SELECT name, employee_name, attendance_device_id, department
           FROM `tabEmployee` WHERE status='Active'
           ORDER BY department, employee_name LIMIT 100""",
        as_dict=True,
    )
    departments = frappe.db.sql_list(
        "SELECT DISTINCT department FROM `tabEmployee` WHERE status='Active' AND department IS NOT NULL"
    )
    return {
        "employees": employees,
        "departments": departments,
        "expected_columns": {
            "employee_col": "Colonne contenant le matricule (attendance_device_id) ou le nom Frappe (HR-EMP-XXXX)",
            "date_col": "Colonne date (format : YYYY-MM-DD ou DD/MM/YYYY)",
            "time_col": "Colonne heure (format : HH:MM ou HH:MM:SS)",
            "log_type_col": "Colonne IN/OUT (optionnel : IN, OUT, E, S, Entrée, Sortie)",
        },
        "example_url": "/api/method/kya_hr.kya_hr.api.import_attendance_excel"
                       "?file_url=/files/pointage.xlsx"
                       "&employee_col=Matricule&date_col=Date&time_col=Heure",
    }


# ═══════════════════════════════════════════════════════════════════════
#  IMPORT ÉVALUATION TRIMESTRIELLE (Excel → Plan Trimestriel + Tâches)
# ═══════════════════════════════════════════════════════════════════════

@frappe.whitelist()
def import_evaluation_excel(file_url=None, equipe=None, trimestre=None, annee=None):
    """
    Importe un fichier Excel d'évaluation trimestrielle collective.
    Crée un Plan Trimestriel + résultats + Tache Equipe pour chaque ligne.

    Format attendu du fichier Excel :
        Col C = N° résultat
        Col D = Résultat attendu
        Col E = Tâche prévue
        Col F = Possibilité de digitalisation (OUI/NON)
        Col G = Taux de digitalisation (0-1 ou %)
        Col H = Attribution (nom employé)
        Col I = Indicateur KPI
        Col J = Taux estimé (0-1 ou %)
        Col K = Taux effectif (0-1 ou %)
        Row 4 = en-tête, données à partir de row 5.
    """
    import openpyxl
    import os

    frappe.only_for(["System Manager", "HR Manager", "HR User", "Chef Service"])

    if not file_url:
        frappe.throw("Paramètre file_url requis (URL du fichier Excel téléchargé)")
    if not equipe:
        frappe.throw("Paramètre equipe requis (nom de l'équipe KYA)")
    if not trimestre or trimestre not in ("T1", "T2", "T3", "T4"):
        frappe.throw("Paramètre trimestre requis (T1, T2, T3 ou T4)")
    annee = cint(annee) or getdate(today()).year

    # Si Chef Service, vérifier qu'il est chef de cette équipe
    roles = frappe.get_roles(frappe.session.user)
    is_admin = any(r in roles for r in ["System Manager", "HR Manager", "HR User"])
    if not is_admin:
        _verify_chef_equipe(equipe)

    # Charger le fichier
    file_path = frappe.get_site_path("public", file_url.lstrip("/"))
    if not os.path.isfile(file_path):
        file_path = frappe.get_site_path(file_url.lstrip("/"))
    if not os.path.isfile(file_path):
        frappe.throw(f"Fichier introuvable : {file_url}")

    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active

    # Parser le titre depuis la feuille (row 2 souvent = "Équipe X – T1 2026")
    titre_plan = ws.cell(row=2, column=3).value or f"Plan {equipe} {trimestre} {annee}"

    # Vérifier s'il existe déjà
    existing = frappe.db.exists("Plan Trimestriel", {
        "equipe": equipe, "trimestre": trimestre, "annee": annee
    })
    if existing:
        frappe.throw(
            f"Un Plan Trimestriel existe déjà pour {equipe} {trimestre} {annee} : {existing}. "
            "Supprimez-le d'abord ou utilisez un autre trimestre."
        )

    # ── Parsing des résultats et tâches ──
    resultats = {}  # {numero: {"libelle": ..., "taches": [...]}}
    current_numero = None

    for row_idx in range(5, ws.max_row + 1):
        c_val = ws.cell(row=row_idx, column=3).value  # N° résultat
        d_val = ws.cell(row=row_idx, column=4).value  # Résultat attendu
        e_val = ws.cell(row=row_idx, column=5).value  # Tâche prévue
        f_val = ws.cell(row=row_idx, column=6).value  # Digitalisable
        g_val = ws.cell(row=row_idx, column=7).value  # Taux digitalisation
        h_val = ws.cell(row=row_idx, column=8).value  # Attribution
        i_val = ws.cell(row=row_idx, column=9).value  # KPI
        j_val = ws.cell(row=row_idx, column=10).value  # Taux estimé
        k_val = ws.cell(row=row_idx, column=11).value  # Taux effectif

        # Nouvelle ligne de résultat
        if c_val is not None:
            try:
                current_numero = int(c_val)
            except (ValueError, TypeError):
                continue
            resultats[current_numero] = {
                "libelle": str(d_val or "").strip(),
                "taches": [],
            }

        # Ligne de tâche (peut être sous un résultat existant)
        if e_val and current_numero and current_numero in resultats:
            tache_data = {
                "libelle": str(e_val).strip(),
                "digitalisable": "OUI" if str(f_val or "").upper().startswith("OUI") else "NON",
                "taux_digitalisation": _parse_pct(g_val),
                "attribution": str(h_val or "").strip() if h_val else None,
                "kpi": str(i_val or "").strip() if i_val else None,
                "taux_estime": _parse_pct(j_val),
                "taux_effectif": _parse_pct(k_val),
            }
            resultats[current_numero]["taches"].append(tache_data)

    if not resultats:
        frappe.throw("Aucun résultat trouvé dans le fichier Excel. "
                     "Vérifiez que les données commencent à la ligne 5, colonne C.")

    # ── Créer le Plan Trimestriel ──
    plan = frappe.get_doc({
        "doctype": "Plan Trimestriel",
        "titre": str(titre_plan).strip(),
        "equipe": equipe,
        "trimestre": trimestre,
        "annee": annee,
        "statut": "En cours",
        "resultats": [],
    })

    # Ajouter les résultats
    for num in sorted(resultats.keys()):
        plan.append("resultats", {
            "numero": num,
            "libelle": resultats[num]["libelle"],
            "poids": round(100.0 / len(resultats), 1),
            "score": 0,
        })

    plan.insert(ignore_permissions=True)
    frappe.db.commit()

    # ── Créer les Tâches par résultat ──
    taches_created = 0
    taches_errors = []

    for num in sorted(resultats.keys()):
        for tache_data in resultats[num]["taches"]:
            try:
                tache = frappe.get_doc({
                    "doctype": "Tache Equipe",
                    "plan": plan.name,
                    "resultat_numero": num,
                    "resultat_libelle": resultats[num]["libelle"][:140],
                    "libelle": tache_data["libelle"],
                    "kpi": tache_data["kpi"],
                    "taux_estime": tache_data["taux_estime"],
                    "taux_effectif": tache_data["taux_effectif"],
                    "digitalisable": tache_data["digitalisable"],
                    "taux_digitalisation": tache_data["taux_digitalisation"],
                    "statut": "Non démarré",
                    "attributions": [],
                })

                # Résoudre l'attribution (nom employé → Employee link)
                if tache_data["attribution"]:
                    emp = _resolve_employee(tache_data["attribution"])
                    if emp:
                        tache.append("attributions", {
                            "employe": emp,
                            "role_attribution": "Responsable",
                        })

                tache.insert(ignore_permissions=True)
                taches_created += 1
            except Exception as e:
                taches_errors.append(f"Résultat {num}, Tâche '{tache_data['libelle'][:40]}': {str(e)}")

    frappe.db.commit()

    return {
        "success": True,
        "plan_name": plan.name,
        "resultats_count": len(resultats),
        "taches_created": taches_created,
        "errors": taches_errors,
        "message": f"Plan {plan.name} créé avec {len(resultats)} résultats et {taches_created} tâches.",
    }


def _parse_pct(val):
    """Convertit une valeur en pourcentage (0-100). Gère 0.5, '50%', 50, '#DIV/0!'."""
    if val is None:
        return 0.0
    if isinstance(val, str):
        val = val.strip().replace("%", "").replace(",", ".")
        if not val or "DIV" in val or "N/A" in val.upper():
            return 0.0
        try:
            val = float(val)
        except ValueError:
            return 0.0
    if isinstance(val, (int, float)):
        # Si entre 0 et 1 (exclusif), convertir en %
        if 0 < val < 1:
            return round(val * 100, 1)
        return round(float(val), 1)
    return 0.0


def _resolve_employee(name_str):
    """Résout un nom d'employé vers un Employee ID Frappe."""
    if not name_str:
        return None
    name_str = name_str.strip()
    # Chercher par employee_name exact
    emp = frappe.db.get_value("Employee", {"employee_name": name_str, "status": "Active"}, "name")
    if emp:
        return emp
    # Chercher par nom partiel (contient)
    emp = frappe.db.get_value(
        "Employee",
        {"employee_name": ["like", f"%{name_str}%"], "status": "Active"},
        "name",
    )
    return emp


@frappe.whitelist()
def get_import_evaluation_template():
    """Retourne la description du format Excel attendu pour l'import."""
    return {
        "format": {
            "row_4": "En-tête (ignoré)",
            "row_5+": "Données",
            "col_C": "N° résultat (entier, marque le début d'un nouveau résultat)",
            "col_D": "Résultat attendu (texte)",
            "col_E": "Tâche prévue (texte, une par ligne)",
            "col_F": "Digitalisable (OUI/NON)",
            "col_G": "Taux de digitalisation (0-1 ou 0-100)",
            "col_H": "Attribution (nom de l'employé)",
            "col_I": "Indicateur KPI (texte)",
            "col_J": "Taux estimé (0-1 ou 0-100)",
            "col_K": "Taux effectif (0-1 ou 0-100, rempli par l'employé)",
        },
        "parameters": {
            "file_url": "/files/evaluation_T1_2026.xlsx (URL du fichier uploadé)",
            "equipe": "Nom du département (ex: Équipe Informatique et Logiciels)",
            "trimestre": "T1, T2, T3 ou T4",
            "annee": "2026 (optionnel, défaut = année en cours)",
        },
        "example": "/api/method/kya_hr.api.import_evaluation_excel"
                   "?file_url=/files/eval.xlsx&equipe=Département IT&trimestre=T1&annee=2026",
    }


# ═══════════════════════════════════════════════════════════════════════
#  ESPACE EMPLOYÉ – MES TÂCHES (consultation + mise à jour taux effectif)
# ═══════════════════════════════════════════════════════════════════════

@frappe.whitelist()
def get_my_tasks(trimestre=None, annee=None):
    """
    Retourne les tâches attribuées à l'employé connecté pour un trimestre.
    L'employé peut voir ses tâches et mettre à jour le taux_effectif.
    """
    user = frappe.session.user
    if user == "Guest":
        return {"tasks": [], "error": "Non connecté"}

    employee = frappe.db.get_value("Employee", {"user_id": user, "status": "Active"}, "name")
    if not employee:
        return {"tasks": [], "error": "Aucun employé actif lié à ce compte"}

    annee = cint(annee) or getdate(today()).year

    filters = {"employe": employee}
    attributions = frappe.get_all(
        "Tache Equipe Attribution",
        filters=filters,
        fields=["parent"],
        limit=500,
    )
    task_names = list({a.parent for a in attributions})
    if not task_names:
        return {"tasks": [], "employee": employee, "trimestre": trimestre, "annee": annee}

    # Filtrer par Plan Trimestriel du bon trimestre
    tache_filters = {"name": ["in", task_names]}
    tasks = frappe.get_all(
        "Tache Equipe",
        filters=tache_filters,
        fields=[
            "name", "plan", "libelle", "kpi", "statut",
            "taux_estime", "taux_effectif", "digitalisable",
            "taux_digitalisation", "resultat_numero", "resultat_libelle",
            "commentaire",
        ],
        order_by="resultat_numero asc",
    )

    # Filtrer côté Python par plan trimestriel (trimestre + année)
    plan_cache = {}
    filtered = []
    for t in tasks:
        if t.plan not in plan_cache:
            plan_cache[t.plan] = frappe.db.get_value(
                "Plan Trimestriel", t.plan,
                ["trimestre", "annee", "equipe", "titre"],
                as_dict=True,
            )
        plan_info = plan_cache.get(t.plan)
        if not plan_info:
            continue
        if trimestre and plan_info.trimestre != trimestre:
            continue
        if plan_info.annee != annee:
            continue
        t["plan_titre"] = plan_info.titre
        t["plan_equipe"] = plan_info.equipe
        t["plan_trimestre"] = plan_info.trimestre
        t["plan_annee"] = plan_info.annee
        filtered.append(t)

    # Stats
    total = len(filtered)
    terminees = sum(1 for t in filtered if t.statut == "Terminé")
    en_cours = sum(1 for t in filtered if t.statut == "En cours")

    taux_moyen = round(
        sum(t.taux_effectif or 0 for t in filtered) / total, 1
    ) if total else 0

    return {
        "tasks": filtered,
        "employee": employee,
        "trimestre": trimestre,
        "annee": annee,
        "stats": {
            "total": total,
            "terminees": terminees,
            "en_cours": en_cours,
            "non_demarrees": total - terminees - en_cours,
            "taux_moyen": taux_moyen,
        },
    }


@frappe.whitelist()
def update_task_progress(task_name, taux_effectif, commentaire=None):
    """
    Permet à l'employé attribué de mettre à jour le taux effectif de sa tâche.
    Le contrôleur Tache Equipe auto-met à jour le statut dans validate().
    """
    user = frappe.session.user
    employee = frappe.db.get_value("Employee", {"user_id": user, "status": "Active"}, "name")
    if not employee:
        frappe.throw("Aucun employé actif lié à ce compte")

    # Vérifier que l'employé est bien attribué à cette tâche
    is_assigned = frappe.db.exists("Tache Equipe Attribution", {
        "parent": task_name,
        "employe": employee,
    })
    if not is_assigned:
        frappe.throw("Vous n'êtes pas attribué à cette tâche", frappe.PermissionError)

    taux = _parse_pct(taux_effectif)
    if taux < 0 or taux > 100:
        frappe.throw("Le taux effectif doit être entre 0 et 100")

    tache = frappe.get_doc("Tache Equipe", task_name)
    tache.taux_effectif = taux
    if commentaire is not None:
        tache.commentaire = str(commentaire)[:500]
    tache.save(ignore_permissions=True)
    frappe.db.commit()

    return {
        "success": True,
        "task_name": task_name,
        "taux_effectif": tache.taux_effectif,
        "statut": tache.statut,
        "message": f"Taux effectif mis à jour : {tache.taux_effectif}%",
    }


# ═══════════════════════════════════════════════════════════════════════
#  GESTION D'ÉQUIPE — APIs scopées par Equipe KYA
# ═══════════════════════════════════════════════════════════════════════

def _get_employee_for_user(user=None):
    """Retourne l'Employee ID de l'utilisateur connecté."""
    user = user or frappe.session.user
    return frappe.db.get_value("Employee", {"user_id": user, "status": "Active"}, "name")


def _get_chef_equipes(employee):
    """Retourne les Equipe KYA dont cet employé est chef."""
    if not employee:
        return []
    return frappe.get_all("Equipe KYA", filters={"chef_equipe": employee, "est_active": 1}, pluck="name")


def _verify_chef_equipe(equipe_name):
    """Vérifie que l'utilisateur connecté est bien le chef de cette équipe. Sinon throw."""
    emp = _get_employee_for_user()
    if not emp:
        frappe.throw("Aucun employé actif lié à ce compte", frappe.PermissionError)
    chef = frappe.db.get_value("Equipe KYA", equipe_name, "chef_equipe")
    if chef != emp:
        frappe.throw(
            f"Vous n'êtes pas le chef de l'équipe '{equipe_name}'.",
            frappe.PermissionError,
        )


@frappe.whitelist()
def get_team_members(equipe=None):
    """
    Retourne les membres d'une équipe.
    Chef Service: ne voit que ses équipes.
    HR/Admin: peut voir toute équipe.
    """
    roles = frappe.get_roles(frappe.session.user)
    is_admin = any(r in roles for r in ["System Manager", "HR Manager", "HR User"])
    emp = _get_employee_for_user()

    if equipe:
        if not is_admin:
            _verify_chef_equipe(equipe)
        equipes = [equipe]
    else:
        if is_admin:
            equipes = frappe.get_all("Equipe KYA", filters={"est_active": 1}, pluck="name")
        else:
            equipes = _get_chef_equipes(emp)
            if not equipes:
                return {"members": [], "error": "Vous n'êtes chef d'aucune équipe"}

    members = []
    for eq in equipes:
        eq_info = frappe.db.get_value("Equipe KYA", eq,
            ["nom_equipe", "departement", "chef_equipe", "chef_equipe_name"], as_dict=True)
        emps = frappe.get_all("Employee",
            filters={"custom_kya_equipe": eq, "status": "Active"},
            fields=["name", "employee_name", "designation", "department", "image",
                     "employment_type", "user_id"],
            order_by="employee_name asc",
        )
        members.append({
            "equipe": eq,
            "equipe_info": eq_info,
            "employees": emps,
            "count": len(emps),
        })

    return {"teams": members, "total_members": sum(m["count"] for m in members)}


@frappe.whitelist()
def get_team_dashboard(equipe=None, trimestre=None, annee=None):
    """
    Dashboard d'équipe pour le Chef d'Équipe.
    Retourne : membres, plans, tâches avec progression, KPIs.
    Le Chef ne voit QUE son équipe.
    """
    roles = frappe.get_roles(frappe.session.user)
    is_admin = any(r in roles for r in ["System Manager", "HR Manager", "HR User"])
    emp = _get_employee_for_user()
    annee = cint(annee) or getdate(today()).year

    # Déterminer l'équipe
    if equipe:
        if not is_admin:
            _verify_chef_equipe(equipe)
    else:
        equipes = _get_chef_equipes(emp) if not is_admin else \
            frappe.get_all("Equipe KYA", filters={"est_active": 1}, pluck="name")
        if not equipes:
            return {"error": "Aucune équipe trouvée"}
        equipe = equipes[0]

    eq_info = frappe.db.get_value("Equipe KYA", equipe,
        ["nom_equipe", "departement", "chef_equipe", "chef_equipe_name", "nombre_membres"],
        as_dict=True)

    # Membres
    membres = frappe.get_all("Employee",
        filters={"custom_kya_equipe": equipe, "status": "Active"},
        fields=["name", "employee_name", "designation", "image"],
    )

    # Plans trimestriels de cette équipe
    plan_filters = {"equipe": equipe, "annee": annee}
    if trimestre:
        plan_filters["trimestre"] = trimestre
    plans = frappe.get_all("Plan Trimestriel",
        filters=plan_filters,
        fields=["name", "titre", "trimestre", "annee", "statut",
                "nombre_taches", "taches_terminees"],
        order_by="trimestre asc",
    )

    # Tâches de cette équipe pour la période
    taches = []
    for plan in plans:
        plan_taches = frappe.get_all("Tache Equipe",
            filters={"plan": plan.name},
            fields=["name", "libelle", "statut", "taux_estime", "taux_effectif",
                    "resultat_numero", "resultat_libelle"],
        )
        for t in plan_taches:
            # Charger les attributions
            t["attributions"] = frappe.get_all("Tache Equipe Attribution",
                filters={"parent": t.name},
                fields=["employe", "nom_employe", "role_attribution"],
            )
            t["plan_name"] = plan.name
            t["plan_titre"] = plan.titre
        taches.extend(plan_taches)

    # Stats agrégées
    total_taches = len(taches)
    terminees = sum(1 for t in taches if t.statut == "Terminé")
    en_cours = sum(1 for t in taches if t.statut == "En cours")
    taux_moyen = round(
        sum(t.taux_effectif or 0 for t in taches) / total_taches, 1
    ) if total_taches else 0

    # Progression par membre
    member_progress = {}
    for m in membres:
        m_tasks = [t for t in taches if any(
            a["employe"] == m.name for a in t.get("attributions", [])
        )]
        m_total = len(m_tasks)
        m_taux = round(
            sum(t.taux_effectif or 0 for t in m_tasks) / m_total, 1
        ) if m_total else 0
        member_progress[m.name] = {
            "employee_name": m.employee_name,
            "designation": m.designation,
            "image": m.image,
            "total_taches": m_total,
            "taux_moyen": m_taux,
            "terminees": sum(1 for t in m_tasks if t.statut == "Terminé"),
        }

    return {
        "equipe": equipe,
        "equipe_info": eq_info,
        "membres": membres,
        "plans": plans,
        "taches": taches,
        "stats": {
            "total_membres": len(membres),
            "total_taches": total_taches,
            "terminees": terminees,
            "en_cours": en_cours,
            "non_demarrees": total_taches - terminees - en_cours,
            "taux_moyen": taux_moyen,
        },
        "member_progress": member_progress,
    }


@frappe.whitelist()
def get_my_equipes():
    """Retourne les équipes de l'utilisateur connecté (en tant que chef ou membre)."""
    emp = _get_employee_for_user()
    if not emp:
        return {"chef_de": [], "membre_de": None}

    chef_de = frappe.get_all("Equipe KYA",
        filters={"chef_equipe": emp, "est_active": 1},
        fields=["name", "nom_equipe", "departement", "nombre_membres"],
    )

    membre_de = frappe.db.get_value("Employee", emp, "custom_kya_equipe")
    equipe_info = None
    if membre_de:
        equipe_info = frappe.db.get_value("Equipe KYA", membre_de,
            ["name", "nom_equipe", "departement", "chef_equipe", "chef_equipe_name"],
            as_dict=True)

    return {
        "employee": emp,
        "chef_de": chef_de,
        "membre_de": equipe_info,
        "is_chef": len(chef_de) > 0,
    }


@frappe.whitelist()
def assign_task_to_member(task_name, employe, role_attribution="Contributeur"):
    """
    Ajouter un membre à une tâche. Seul le chef de l'équipe peut le faire.
    """
    emp = _get_employee_for_user()
    tache = frappe.get_doc("Tache Equipe", task_name)

    # Vérifier que le chef est bien chef de l'équipe de cette tâche
    equipe = tache.equipe
    roles = frappe.get_roles(frappe.session.user)
    is_admin = any(r in roles for r in ["System Manager", "HR Manager"])
    if not is_admin:
        _verify_chef_equipe(equipe)

    # Vérifier que l'employé ciblé est bien membre de cette équipe
    emp_equipe = frappe.db.get_value("Employee", employe, "custom_kya_equipe")
    if emp_equipe != equipe:
        frappe.throw(
            f"L'employé {employe} n'appartient pas à l'équipe '{equipe}'.",
            frappe.ValidationError,
        )

    # Vérifier pas déjà attribué
    already = any(a.employe == employe for a in tache.attributions)
    if already:
        frappe.throw(f"L'employé {employe} est déjà attribué à cette tâche.")

    tache.append("attributions", {
        "employe": employe,
        "role_attribution": role_attribution if role_attribution in ("Responsable", "Contributeur") else "Contributeur",
    })
    tache.save(ignore_permissions=True)
    frappe.db.commit()

    emp_name = frappe.db.get_value("Employee", employe, "employee_name")
    return {
        "success": True,
        "message": f"{emp_name} ajouté(e) à la tâche '{tache.libelle}'",
    }
